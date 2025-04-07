import streamlit as st
from langchain_community.chat_models import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from langchain_community.document_loaders import TextLoader, Docx2txtLoader, UnstructuredPowerPointLoader
from langchain.text_splitter import CharacterTextSplitter
from langchain_community.embeddings import OpenAIEmbeddings
from langchain_community.vectorstores import FAISS
import tempfile
import os
import openpyxl
from langchain_community.document_loaders.base import BaseLoader
from langchain.docstore.document import Document

# Set up OpenAI API key
os.environ["OPENAI_API_KEY"] = "your-api-key-here"

# Custom Excel Loader
class ExcelLoader(BaseLoader):
    def __init__(self, file_path):
        self.file_path = file_path

    def load(self):
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        documents = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content = f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                content += " | ".join(str(cell) for cell in row if cell is not None) + "\n"
            documents.append(Document(page_content=content, metadata={"source": self.file_path, "sheet": sheet_name}))
        return documents

# Function to load and process the uploaded file
def process_file(uploaded_file):
    # Create a temporary file to store the uploaded content
    with tempfile.NamedTemporaryFile(delete=False, suffix=f".{uploaded_file.name.split('.')[-1]}") as temp_file:
        temp_file.write(uploaded_file.getvalue())
        temp_file_path = temp_file.name

    # Determine the file type and use the appropriate loader
    file_extension = uploaded_file.name.split('.')[-1].lower()
    
    if file_extension == 'txt':
        loader = TextLoader(temp_file_path)
    elif file_extension == 'docx':
        loader = Docx2txtLoader(temp_file_path)
    elif file_extension == 'xlsx':
        loader = ExcelLoader(temp_file_path)
    elif file_extension == 'pptx':
        loader = UnstructuredPowerPointLoader(temp_file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_extension}")

    documents = loader.load()
    
    # Split the documents into chunks
    text_splitter = CharacterTextSplitter(chunk_size=1000, chunk_overlap=0)
    texts = text_splitter.split_documents(documents)

    # Create embeddings and store them in a vector database
    embeddings = OpenAIEmbeddings()
    vectorstore = FAISS.from_documents(texts, embeddings)

    # Clean up the temporary file
    os.unlink(temp_file_path)

    return vectorstore

# Set up the Streamlit interface
st.title("File Upload Chatbot")

# File uploader
uploaded_file = st.file_uploader("Choose a file", type=["txt", "docx", "xlsx", "pptx"])

if uploaded_file is not None:
    vectorstore = process_file(uploaded_file)

    # Set up the conversational chain
    memory = ConversationBufferMemory(memory_key="chat_history", return_messages=True)
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=ChatOpenAI(temperature=0.7),
        retriever=vectorstore.as_retriever(),
        memory=memory
    )

    # Chat interface
    if "messages" not in st.session_state:
        st.session_state.messages = []

    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    if prompt := st.chat_input("What would you like to know about the file?"):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        with st.chat_message("assistant"):
            response = conversation_chain({"question": prompt})
            st.markdown(response['answer'])
        st.session_state.messages.append({"role": "assistant", "content": response['answer']})

else:
    st.write("Please upload a file to start the conversation.")